# 載入需要的套件
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from time import sleep
from bs4 import BeautifulSoup
import pandas as pd
from urllib.parse import urljoin
import time
from openpyxl import load_workbook
import openpyxl
import logging
import os
import xlsxwriter

#設定 log 的 filename 後只會輸出到檔案不會輸出在 console
logging.basicConfig(level=logging.INFO, filename='./log.txt', filemode='a',
    format='[%(asctime)s %(levelname)s] %(message)s',
    datefmt='%Y%m%d %H:%M:%S',
    )

def scroll(driver,chrome_options):
    # driver = webdriver.Chrome(options = chrome_options)
    # driver.get(url)
    SCROLL_PAUSE_TIME = 1
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
    source = driver.page_source
    driver.close()
    return source

def open_csv(filename):
    professor_name_list=[]
    file = load_workbook(filename)
    # 取得第一個工作表
    sheet = file.worksheets[0]
    # 顯示 row總數 及 column總數
    print('row總數:', sheet.max_row)
    print('column總數:', sheet.max_column)
    # 顯示 cell 資料
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            professor_name_list.append(sheet.cell(row = i, column = j).value)
    print("總共有 "+str(len(professor_name_list))+" 個教授")
    return professor_name_list

def write_to_csv(df,professor_name,department):
    # 取得目前工作目錄路徑
    path = os.getcwd()
    filename = professor_name + '.xlsx'
    filepath = os.path.join(path,department, filename)
    # df.to_excel(filepath, sheet_name=professor_name)
    df.to_excel(filepath, sheet_name=professor_name, engine='xlsxwriter')

def crawler_NDLTD(professor_name,department):
    while True:
        try:
            start = time.time()
            # 開啟瀏覽器視窗(Chrome)
            chrome_options = Options()
            chrome_options.add_argument('--headless=new')
            driver = webdriver.Chrome(options=chrome_options)
            print('Get New Cookie')
            driver.get('https://ndltd.ncl.edu.tw/cgi-bin/gs32/gsweb.cgi?o=d')
            sleep(2)
            # driver.find_element_by_xpath('//a[@title="進階查詢"]').click()
            element = driver.find_element(By.XPATH, '//a[@title="進階查詢"]')
            element.click()
            sleep(2)
            driver.find_element(By.ID,'ysearchinput0').send_keys(professor_name)
            # 找到下拉選單元素
            dropdown_element = driver.find_element(By.NAME,"qf0")
            # 創建一個 Select 物件
            select = Select(dropdown_element)
            # 或通過選項的值選擇
            select.select_by_value("ad")
            sleep(0.5)
            # 搜尋
            driver.find_element(By.ID,'gs32search').click()
            sleep(2)  
            break
        except:
            print("異常發生ㄌ")
    # scroll(driver,chrome_options)
    dictionary = ["論文永久網址:","研究生:","研究生(外文):","論文名稱:","論文名稱(外文):","指導教授:","指導教授(外文):","學位類別:","校院名稱:","系所名稱:","學門:",\
                  "學類:","論文種類:","論文出版年:","畢業學年度:","語文別:","論文頁數:","中文關鍵詞:","外文關鍵詞:","中文摘要","英文摘要"]
    df = pd.DataFrame(columns=dictionary)
    current_row =1
    current_page =0
    print(f"開始爬取 "+professor_name+" 資料～")
    data_num=100

    while current_page*20+current_row < data_num:

        # 使用 CSS 選擇器選擇特定的超連結
        if current_row%20==1 and current_row!=1:
            try:
                # 獲取網頁高度
                page_height = driver.execute_script("return document.body.scrollHeight")
                # 持續滾動網頁直到底部
                while True:
                    # 滾動至底部
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    # 等待一些時間，讓新的內容加載
                    driver.implicitly_wait(2)  # 等待 2 秒
                    # 檢查是否已經滾動到底部
                    new_page_height = driver.execute_script("return document.body.scrollHeight")
                    if new_page_height == page_height:
                        break
                    page_height = new_page_height
            except Exception as e:
                print("發生錯誤:", e)
            # 使用 XPath 定位下一頁按鈕元素
            next_page_input = driver.find_element(By.XPATH, '//td/input[@type="image" and @alt="下一頁"]')
            # 點擊下一頁的 input 元素
            next_page_input.click()
            
        record='a.slink[href*="/record?r1='+str(current_row)+'&h1='+str(current_page)+'"] span.etd_d'
        try:
            link_element = driver.find_element(By.CSS_SELECTOR, record)
            link_element.click()
        except NoSuchElementException:
            # 创建一个新的Excel工作簿
            workbook = openpyxl.Workbook()
            # 保存工作簿到文件
            filename=professor_name+'.xlsx'
            filepath = os.path.join(path,department, filename)
            workbook.save(filepath)
            # 关闭工作簿（可选）
            workbook.close()
            end = time.time()
            logging.info("空白 save to "+department+" "+professor_name+".xlsx spend "+str((end - start)/60)+" 分鐘")
            return
        page_source = driver.page_source
        # 使用 Beautiful Soup 解析源代碼
        soup = BeautifulSoup(page_source, 'html.parser')

        if current_row==1 and current_page==0:
            # 找到包含目標文字的<label>標籤
            label_text = soup.find('label', {'for': 'browsechecker'}).get_text(strip=True)
            l=label_text.split(" ")
            data_num=int(l[10])
            print("總共 "+str(data_num)+" 筆")
            data_num = data_num+1

        print(f"開始爬取第{current_page}頁的第{current_row}筆")

        meta_list=[""]*21
        for num,table in enumerate(soup.find_all('table', {'class': 'tableoutfmt2'})):
            if num==0:
                for i in table.findAll('tr'):
                    # 論文永久網址
                    if i.select_one('input.pushurlcls1') is not None:
                        meta_list[0]=i.select_one('input.pushurlcls1')["value"]
                    # 研究生中文名
                    if i.select_one('th[id="format_0_table_th_1"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_1"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 研究生英文名
                    if i.select_one('th[id="format_0_table_th_2"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_2"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 論文名稱
                    if i.select_one('th[id="format_0_table_th_3"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_3"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 論文英文名稱
                    if i.select_one('th[id="format_0_table_th_4"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_4"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 指導教授
                    if i.select_one('th[id="format_0_table_th_5"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_5"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                             # print("找不到指定值在字典中。")
                            pass
                    # 指導教授（外文）
                    if i.select_one('th[id="format_0_table_th_6"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_6"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 學位類別
                    if i.select_one('th[id="format_0_table_th_7"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_7"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 校院名稱
                    if i.select_one('th[id="format_0_table_th_8"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_8"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 系所名稱
                    if i.select_one('th[id="format_0_table_th_9"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_9"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 學門
                    if i.select_one('th[id="format_0_table_th_10"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_10"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 學類
                    if i.select_one('th[id="format_0_table_th_11"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_11"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 論文種類
                    if i.select_one('th[id="format_0_table_th_12"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_12"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 論文出版年
                    if i.select_one('th[id="format_0_table_th_13"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_13"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 畢業學年度
                    if i.select_one('th[id="format_0_table_th_14"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_14"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 語文別
                    if i.select_one('th[id="format_0_table_th_15"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_15"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 頁數
                    if i.select_one('th[id="format_0_table_th_16"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_16"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                keyword.rstrip("、")
                                meta_list[position]=keyword
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        keyword += j.text+'、'
                                    keyword.rstrip("、")
                                meta_list[position]=keyword
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass
                    # 中文關鍵詞
                    if i.select_one('th[id="format_0_table_th_17"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_17"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                            # print("找不到指定值在字典中。")
                            pass

                    # 外文關鍵詞
                    if i.select_one('th[id="format_0_table_th_18"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_18"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                             # print("找不到指定值在字典中。")
                            pass
                    # 緩衝
                    if i.select_one('th[id="format_0_table_th_19"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_19"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                             # print("找不到指定值在字典中。")
                            pass
                    # 緩衝
                    if i.select_one('th[id="format_0_table_th_20"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_20"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                             # print("找不到指定值在字典中。")
                            pass
                    # 緩衝
                    if i.select_one('th[id="format_0_table_th_21"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_21"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                             # print("找不到指定值在字典中。")
                            pass
                    # 緩衝
                    if i.select_one('th[id="format_0_table_th_22"]') is not None:
                        try:
                            position = dictionary.index(i.select_one('th[id="format_0_table_th_22"]').text)
                            if position==1 or position==2 or position==8 or position==9:
                                meta_list[position]=i.select_one('a.slink').text
                            elif position==17 or position==18:
                                keyword=""
                                for j in i.select_one('td.std2').find_all('a'):
                                    keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            elif position==5 or position==6:
                                keyword=""
                                if len(i.select_one('td.std2').find_all('a')) == 0:
                                    keyword = i.select_one('a.slink').text
                                else:
                                    for j in i.select_one('td.std2').find_all('a'):
                                        if j.text!="":
                                            keyword += j.text+'、'
                                meta_list[position]=keyword.rstrip("、")
                            else:
                                meta_list[position]=i.select_one('td.std2').text
                        except ValueError:
                             # print("找不到指定值在字典中。")
                            pass
            elif num==1:
                try:
                # 在<table>內部找到<tr>元素
                    for tr in table.findAll('tr'):
                        td = tr.select_one('td.stdncl2')
                        meta_list[19]=td.text
                except:
                    pass
            elif num==2:
                try:
                    # 在<table>內部找到<tr>元素
                    for tr in table.findAll('tr'):
                        td = tr.select_one('td.stdncl2')
                        meta_list[20]=td.text
                except:
                    pass
        df_metadata = pd.DataFrame([meta_list], columns=df.columns)
        df = pd.concat([df_metadata,df],axis=0, ignore_index=True)
        time.sleep(2)
        driver.back()
        time.sleep(5)
        current_row+=1
    write_to_csv(df,professor_name,department)
    end = time.time()
    logging.info("save to "+department+" "+professor_name+".xlsx spend "+str((end - start)/60)+" 分鐘")

if __name__=="__main__":
    # 取得目前工作目錄路徑
    path = os.getcwd()
    department="自動化"
    filename = '教授名稱.xlsx'
    filepath = os.path.join(path,department, filename)
    professor_name_list=open_csv(filepath)
    
    for name in professor_name_list[1:]:
        filename = name + '.xlsx'
        filepath = os.path.join(path,department, filename)
        if os.path.exists(filepath):
            print(name+" have saved to " + filepath)
            continue
        crawler_NDLTD(name,department)
    logging.info("自動化學門結束囉～～")  

    department="控制"
    filename = '教授名稱.xlsx'
    filepath = os.path.join(path,department, filename)
    professor_name_list=open_csv(filepath)
    
    for name in professor_name_list[1:]:
        filename = name + '.xlsx'
        filepath = os.path.join(path,department, filename)
        if os.path.exists(filepath):
            print(name+" have saved to " + filepath)
            continue
        crawler_NDLTD(name,department)
    logging.info("控制學門結束囉～～")  